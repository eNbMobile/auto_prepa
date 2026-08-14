#!/usr/bin/env python3
"""
Workflow "vav" : pour chaque commande listee dans Drive GITHUB/Avoir/
"Commandes en cours" (civilite, nom, prenom, date, creneau), verifie si le
client a un avoir jaune (du, pas encore deduit) dans le classeur Google
Sheets "Avoir/Demarque" — un avoir etant valable 3 mois a compter de sa
DATE DEBUT, les onglets du mois courant et des 3 mois precedents sont
regardes (cf. _onglets_a_verifier). Envoie un mail par avoir trouve, puis
vide "Commandes en cours" (hors en-tete) pour ne pas retraiter ces
commandes au prochain declenchement.
"""
import base64
import io
import os
import sys
import unicodedata
from datetime import date, datetime
from email.mime.text import MIMEText
from zoneinfo import ZoneInfo

import controle_stocks as cs

_TZ = ZoneInfo("Europe/Paris")

MOIS_FR = ["JANVIER", "FEVRIER", "MARS", "AVRIL", "MAI", "JUIN", "JUILLET",
           "AOUT", "SEPTEMBRE", "OCTOBRE", "NOVEMBRE", "DECEMBRE"]

# ID du classeur "AVOIR DEMARQUE DRIVE 2026" (docs.google.com/spreadsheets/d/<id>).
# Utilise si absent de config.json (cle "avoir_spreadsheet_id") — a mettre a
# jour dans config.json les annees suivantes, quand un nouveau classeur est cree.
AVOIR_SPREADSHEET_ID_DEFAUT = "1TEz5Bmx_sIOOSdAoBGB9KXuZAzP3TdhnPCfVDS4yyrs"


def _normaliser(texte):
    """Majuscules sans accents, pour comparer noms/mois insensiblement a la casse/accents."""
    d = unicodedata.normalize('NFD', (texte or '').strip().upper())
    return ''.join(c for c in d if unicodedata.category(c) != 'Mn')


def _initiale(nom_normalise):
    """Premiere lettre alphabetique d'un nom deja normalise ('' si aucune)."""
    for c in nom_normalise:
        if c.isalpha():
            return c
    return ''


def _est_initiale(nom_normalise):
    """Vrai si le nom normalise ne contient qu'une seule lettre (ex. 'R', 'R.')."""
    return sum(c.isalpha() for c in nom_normalise) == 1


def _champs_correspondent(a, b):
    """Compare un nom ou prenom entre 'Commandes en cours' et l'Avoir, en
    ignorant casse/accents (_normaliser) et en tolerant qu'un des deux cotes
    ne soit renseigne que par une initiale (ex. 'R' vs 'RENAUD') — les deux
    fichiers n'etant pas forcement remplis avec le meme niveau de detail."""
    a, b = _normaliser(a), _normaliser(b)
    if not a or not b:
        return False
    if a == b:
        return True
    if _est_initiale(a) or _est_initiale(b):
        return _initiale(a) == _initiale(b)
    return False


# Couleurs de fond reelles du classeur Avoir/Demarque (verifiees sur le
# classeur de production) : jaune plein #FFFF00 = avoir du, vert #00B050 =
# avoir deja deduit. Comparaison a tolerance fine pour absorber les
# arrondis flottants de l'API Sheets, sans dependre d'une heuristique large
# qui matcherait d'autres teintes (orange du total, etc.).
_JAUNE_RGB = (1.0, 1.0, 0.0)
_TOLERANCE = 0.06


def _proche(rgb, cible):
    return all(abs(c - r) < _TOLERANCE for c, r in zip(rgb, cible))


def _est_jaune(rgb):
    """Vrai si le fond de la cellule correspond au jaune #FFFF00 (avoir du)."""
    return _proche(rgb, _JAUNE_RGB)


def _civilite_longue(civilite):
    c = _normaliser(civilite)
    return 'Madame' if c.startswith('MME') or c.startswith('MADAME') else 'Monsieur'


def _formatter_montant(valeur):
    """Formate un montant (nombre lu en cellule xlsx) en '7,98 €'."""
    try:
        return f"{float(valeur):.2f}".replace('.', ',') + " €"
    except (TypeError, ValueError):
        return f"{valeur} €"


def _formatter_date(valeur):
    """Formate une date (datetime lue en cellule xlsx) en 'JJ/MM/AAAA'."""
    if isinstance(valeur, datetime):
        return valeur.strftime('%d/%m/%Y')
    if isinstance(valeur, date):
        return valeur.strftime('%d/%m/%Y')
    return str(valeur)


def _onglets_a_verifier(date_ref, nb_mois=4):
    """Noms d'onglets (normalises) des nb_mois derniers mois, mois courant
    inclus. Un avoir est valable 3 mois a compter de sa DATE DEBUT ; un avoir
    saisi en toute fin de mois (ex. 31 mai) reste donc valide jusqu'a fin du
    4e mois (31 aout) — d'ou nb_mois=4 pour ne rater aucun onglet contenant
    encore des avoirs valides. Le filtre precis reste _date_fin_valide()."""
    onglets = []
    y, m = date_ref.year, date_ref.month
    for i in range(nb_mois):
        mm = m - i
        while mm <= 0:
            mm += 12
        onglets.append(MOIS_FR[mm - 1])
    return onglets


def _date_fin_valide(date_fin, aujourdhui):
    if isinstance(date_fin, datetime):
        date_fin = date_fin.date()
    if isinstance(date_fin, date):
        return date_fin >= aujourdhui
    return True  # type inattendu : on ne filtre pas plutot que de perdre un avoir valide


def _get_sheets_service():
    try:
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request
        from googleapiclient.discovery import build
        if not os.path.exists(cs.TOKEN_FILE):
            return None
        creds = Credentials.from_authorized_user_file(cs.TOKEN_FILE, cs.SCOPES)
        if creds.expired and creds.refresh_token:
            creds.refresh(Request())
        return build("sheets", "v4", credentials=creds)
    except Exception as e:
        print(f"  Sheets inaccessible : {e}")
        return None


def _charger_config_avoir(drive_svc):
    """Retourne (avoir_spreadsheet_id, email_destinataire) depuis config.json."""
    import json
    from googleapiclient.http import MediaIoBaseDownload
    res = drive_svc.files().list(
        q=f"name='config.json' and '{cs.DRIVE_CONFIG_FOLDER_ID}' in parents and trashed=false",
        fields="files(id)",
    ).execute()
    files = res.get("files", [])
    if not files:
        print("ERREUR : config.json introuvable dans le dossier Drive config.")
        sys.exit(1)
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, drive_svc.files().get_media(fileId=files[0]["id"]))
    done = False
    while not done:
        _, done = dl.next_chunk()
    cfg = json.loads(buf.getvalue().decode())
    avoir_id = cfg.get("avoir_spreadsheet_id", "").strip() or AVOIR_SPREADSHEET_ID_DEFAUT
    email = cfg.get("email_destinataire", "").strip()
    if not email:
        print("ERREUR : email_destinataire absent de config.json.")
        sys.exit(1)
    return avoir_id, email


def _argb_vers_rgb01(argb):
    """Convertit une couleur ARGB hex openpyxl (ex. 'FFFFFF00') en tuple RGB
    0-1. None/format inattendu -> blanc (pas de remplissage)."""
    if not argb or len(argb) < 6:
        return (1.0, 1.0, 1.0)
    hexe = argb[-6:]
    try:
        return (int(hexe[0:2], 16) / 255, int(hexe[2:4], 16) / 255, int(hexe[4:6], 16) / 255)
    except ValueError:
        return (1.0, 1.0, 1.0)


def _couleur_cellule(cell):
    """Couleur de fond effective d'une cellule openpyxl (blanc si pas de
    remplissage uni ou couleur de theme non resolue)."""
    fill = cell.fill
    if not fill or fill.patternType != 'solid' or not fill.fgColor:
        return (1.0, 1.0, 1.0)
    fg = fill.fgColor
    if getattr(fg, 'type', None) != 'rgb' or not fg.rgb:
        return (1.0, 1.0, 1.0)
    return _argb_vers_rgb01(fg.rgb)


def _valeur_str(v):
    return str(v).strip() if v not in (None, "") else ""


def _telecharger_avoir_xlsx(drive_svc, spreadsheet_id, chemin):
    """Exporte le classeur Avoir en xlsx (Drive export). Necessaire pour lire
    la couleur de fond reellement affichee : le jaune/vert de ce classeur est
    pose par mise en forme conditionnelle, et effectiveFormat.backgroundColor
    de l'API Sheets ne la reflete pas (verifie empiriquement : renvoie le
    remplissage statique sous-jacent, blanc, pas la couleur conditionnelle
    visible a l'ecran) — l'export xlsx, lui, fige la couleur conditionnelle
    resolue en remplissage statique."""
    from googleapiclient.http import MediaIoBaseDownload
    req = drive_svc.files().export_media(
        fileId=spreadsheet_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    with open(chemin, "wb") as f:
        f.write(buf.getvalue())


def _lire_avoirs_jaunes(drive_svc, spreadsheet_id, date_ref):
    """Lit les lignes a fond jaune (avoir du) des onglets des derniers mois
    (cf. _onglets_a_verifier), via un export xlsx du classeur (cf.
    _telecharger_avoir_xlsx)."""
    import openpyxl

    onglets_cibles = _onglets_a_verifier(date_ref)
    chemin = "_tmp_avoir_demarque.xlsx"
    _telecharger_avoir_xlsx(drive_svc, spreadsheet_id, chemin)
    try:
        wb = openpyxl.load_workbook(chemin, data_only=True)
    finally:
        if os.path.exists(chemin):
            os.remove(chemin)

    correspondance = {}
    for titre in wb.sheetnames:
        norm = _normaliser(titre)
        if norm in onglets_cibles and norm not in correspondance:
            correspondance[norm] = titre
    manquants = [o for o in onglets_cibles if o not in correspondance]
    if manquants:
        print(f"  Onglets Avoir introuvables (ignores) : {', '.join(manquants)}")

    avoirs = []
    for norm in onglets_cibles:
        titre = correspondance.get(norm)
        if not titre:
            continue
        ws = wb[titre]
        for row in ws.iter_rows(min_row=3, max_col=6):
            nom = _valeur_str(row[0].value)
            if not nom:
                continue
            rgb = _couleur_cellule(row[0])
            print(f"    [{titre}] {nom} {_valeur_str(row[1].value)} — fond rgb={rgb}")
            if not _est_jaune(rgb):
                continue
            avoirs.append({
                "onglet": titre, "nom": nom, "prenom": _valeur_str(row[1].value),
                "montant": row[2].value, "date_debut": row[3].value, "date_fin": row[4].value,
                "raison": _valeur_str(row[5].value),
            })
    return avoirs


def _trouver_fichier_commandes(drive_svc):
    """Cherche GITHUB/Avoir/Commandes en cours sur Drive. Retourne (id, mimeType) ou None."""
    res = drive_svc.files().list(
        q=("name='GITHUB' and 'root' in parents "
           "and mimeType='application/vnd.google-apps.folder' and trashed=false"),
        fields="files(id)",
    ).execute()
    github = res.get("files", [])
    if not github:
        print("  Dossier GITHUB/ introuvable sur Drive.")
        return None
    res = drive_svc.files().list(
        q=(f"name='Avoir' and '{github[0]['id']}' in parents "
           f"and mimeType='application/vnd.google-apps.folder' and trashed=false"),
        fields="files(id)",
    ).execute()
    avoir_folder = res.get("files", [])
    if not avoir_folder:
        print("  Dossier GITHUB/Avoir/ introuvable sur Drive.")
        return None
    res = drive_svc.files().list(
        q=(f"name contains 'Commandes en cours' and '{avoir_folder[0]['id']}' in parents "
           f"and trashed=false"),
        fields="files(id,name,mimeType)",
    ).execute()
    fichiers = res.get("files", [])
    if not fichiers:
        print("  'Commandes en cours' introuvable dans GITHUB/Avoir/.")
        return None
    return fichiers[0]["id"], fichiers[0]["mimeType"]


def _lire_commandes_en_cours(sheets_svc, file_id):
    """Retourne [(civilite, nom, prenom, date, creneau), ...] (lignes non vides, hors en-tete)."""
    res = sheets_svc.spreadsheets().values().get(
        spreadsheetId=file_id, range="A2:E1000").execute()
    lignes = []
    for row in res.get("values", []):
        row = row + [""] * (5 - len(row))
        if row[1].strip() or row[2].strip():
            lignes.append(tuple(c.strip() for c in row[:5]))
    return lignes


def _vider_commandes_en_cours(sheets_svc, file_id):
    """Vide les lignes de donnees (garde l'en-tete) pour ne pas retraiter ces
    commandes au prochain declenchement."""
    sheets_svc.spreadsheets().values().clear(
        spreadsheetId=file_id, range="A2:E1000", body={}).execute()
    print("  'Commandes en cours' vide.")


def _envoyer_email_avoir(gmail_svc, destinataire, civilite, avoir, date_cde, creneau):
    if not gmail_svc:
        print("  Gmail inaccessible — email non envoye.")
        return
    civ = _civilite_longue(civilite)
    montant = _formatter_montant(avoir["montant"])
    corps = (
        f"Bonjour,\n\n\n"
        f"Un avoir est dû à {civ} {avoir['nom']} {avoir['prenom']} d'un montant de {montant} "
        f"pour le {date_cde} entre {creneau}\n\n\n"
        f"Cordialement,\n"
        f"Erwan"
    )
    msg = MIMEText(corps, 'plain', 'utf-8')
    msg['To'] = destinataire
    msg['Subject'] = "Avoir à déduire"
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    try:
        gmail_svc.users().messages().send(userId='me', body={'raw': raw}).execute()
        print(f"  Email envoye -> {destinataire} ({avoir['nom']} {avoir['prenom']}, {montant})")
    except Exception as e:
        print(f"  Email echoue ({avoir['nom']} {avoir['prenom']}) : {e}")


def main():
    if not cs.DRIVE_CONFIG_FOLDER_ID:
        print("ERREUR : secret DRIVE_CONFIG_FOLDER_ID manquant.")
        sys.exit(1)

    drive_svc = cs._get_drive_service()
    if not drive_svc:
        print("ERREUR : Drive inaccessible.")
        sys.exit(1)
    sheets_svc = _get_sheets_service()
    if not sheets_svc:
        print("ERREUR : Sheets inaccessible.")
        sys.exit(1)

    avoir_spreadsheet_id, email_destinataire = _charger_config_avoir(drive_svc)

    aujourdhui = datetime.now(_TZ).date()
    print(f"Lecture des avoirs jaunes ({aujourdhui.strftime('%d/%m/%Y')}) …")
    avoirs_jaunes = _lire_avoirs_jaunes(drive_svc, avoir_spreadsheet_id, aujourdhui)
    print(f"→ {len(avoirs_jaunes)} avoir(s) jaune(s) (du) sur les onglets "
          f"{', '.join(_onglets_a_verifier(aujourdhui))}.")

    fichier = _trouver_fichier_commandes(drive_svc)
    if not fichier:
        print("Rien a verifier : 'Commandes en cours' introuvable.")
        return
    file_id, mime_type = fichier
    if mime_type != "application/vnd.google-apps.spreadsheet":
        print(f"ERREUR : 'Commandes en cours' doit etre une Google Sheet (type actuel : {mime_type}).")
        sys.exit(1)

    commandes = _lire_commandes_en_cours(sheets_svc, file_id)
    print(f"{len(commandes)} commande(s) a verifier.")
    if not commandes:
        return

    gmail_svc = cs._get_gmail_service()
    nb_envoyes = 0
    for civilite, nom, prenom, date_cde, creneau in commandes:
        print(f"\nVerification : {civilite} {nom} {prenom} (commande du {date_cde}, {creneau})")
        trouve = False
        for avoir in avoirs_jaunes:
            if not (_champs_correspondent(nom, avoir["nom"])
                    and _champs_correspondent(prenom, avoir["prenom"])):
                continue
            trouve = True
            if not _date_fin_valide(avoir["date_fin"], aujourdhui):
                print(f"  Avoir trouve ({avoir['onglet']}, {avoir['nom']} {avoir['prenom']}) "
                      f"mais expire (date fin {_formatter_date(avoir['date_fin'])}) — ignore.")
                continue
            print(f"  Avoir trouve ({avoir['onglet']}, {avoir['nom']} {avoir['prenom']}, "
                  f"{_formatter_montant(avoir['montant'])}) — envoi du mail.")
            _envoyer_email_avoir(gmail_svc, email_destinataire, civilite, avoir, date_cde, creneau)
            nb_envoyes += 1
        if not trouve:
            print("  Aucun avoir jaune correspondant.")

    print(f"\n{nb_envoyes} email(s) envoye(s).")

    print("\nVidage de 'Commandes en cours' …")
    _vider_commandes_en_cours(sheets_svc, file_id)


if __name__ == "__main__":
    main()
